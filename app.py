import datetime
import openpyxl
import pyperclip
import pyautogui
from time import sleep
import fix_txt
import set_icms

# "Importar" para importar os arquivos para o Sedif
# "AdicionarST" para adicionar os valores de ST aos arquivos
# "GerarProsft" para gerar os pela Prosoft

SelecionarFunção = "AdicionarST"

ComecarApartirDaEmpresa = "0"

MesApuração = '02'
AnoDeApuração = '2024'
CaminhoDaPasta = f"/caminho/da/pasta{MesApuração + AnoDeApuração}"

workbook = openpyxl.load_workbook('planilha.xlsx')
sheet_estado = workbook["Sedif"]

def formatarNumeroEmpresa(Numero):
    NumeroParaString = str(Numero)
    NumeroFormatado = NumeroParaString.zfill(4)
    return NumeroFormatado

def formatarNomeDoArquivo(AnoDeApuração, MesApuração, NumeroFormatado):
    # Formato do arquivo final DeSTDA_202311-0566.txt
    NomeDoArquivo = 'DeSTDA_' + AnoDeApuração + MesApuração + '-' + NumeroFormatado + '.txt'
    return NomeDoArquivo


def GerarArquivoProsoft():
    sleep(6)
    for linha in sheet_estado.iter_rows(min_row=2):
        # Prosoft
        SelecionarEmpresa = (751,216);
        SelecionarMes = (554,242);
        SelecionarPasta = (1153,589);
        SalvarArquivo = (957,530);
        GerarTxt = (887,661);
        DarOK = (882,507);
        
        NumeroFormatado = formatarNumeroEmpresa(linha[0].value)
        NomeDoArquivo = formatarNomeDoArquivo(AnoDeApuração, MesApuração, NumeroFormatado)

        print(NomeDoArquivo)
        if NumeroFormatado >= ComecarApartirDaEmpresa:
            pyperclip.copy(NumeroFormatado)
            pyautogui.click(SelecionarEmpresa,duration=1)
            pyautogui.hotkey('ctrl','v')
            pyautogui.press('enter')
            sleep(0.5)
            pyperclip.copy(MesApuração)
            pyautogui.hotkey('ctrl','v')
            pyautogui.press('enter')
            sleep(0.5)
            pyautogui.click(SelecionarPasta,duration=1)
            sleep(0.5)
            pyperclip.copy(NomeDoArquivo)
            pyautogui.hotkey('ctrl','v')
            sleep(0.2)
            pyautogui.press('enter')
            sleep(0.5)
            pyautogui.click(GerarTxt,duration=1)
            sleep(1)
            pyautogui.click(DarOK,duration=1)

def ImportarArquivoSedif():
    sleep(6)
    ImportarArquivo = (375,85);
    SelecionarArquivo = (1578,137);
    Importar = (75,165);
    Confirmação = (868,510);
    AvisoDuplicado = (855,555);
    Aviso = (886,503);
    for linha in sheet_estado.iter_rows(min_row=2):
        
        if linha[0].value == None:
            break
        # Sedif
        NumeroFormatado = formatarNumeroEmpresa(linha[0].value)
        NomeDoArquivo = formatarNomeDoArquivo(AnoDeApuração, MesApuração, NumeroFormatado)
        print(NomeDoArquivo)
        if NumeroFormatado >= ComecarApartirDaEmpresa:
            print("Começou a importar apartir da empresa: " + ComecarApartirDaEmpresa)
            pyautogui.click(ImportarArquivo,duration=1)
            sleep(1)
            pyperclip.copy(NomeDoArquivo)
            pyautogui.click(SelecionarArquivo,duration=1)
            sleep(1)
            pyautogui.hotkey('ctrl','v')
            pyautogui.press('enter')
            sleep(1)
            pyautogui.click(Importar,duration=1)
            sleep(1.5)
            pyautogui.click(Confirmação,duration=1)
            sleep(1.5)
            # pyautogui.click(AvisoDuplicado,duration=1)
            # sleep(5)
            pyautogui.click(Aviso,duration=1)
            sleep(1)
        else:
            print("Empresa: " + NumeroFormatado + " foi pulada")

def AdicionarST():
    for linha in sheet_estado.iter_rows(min_row=2):
        if linha[1].value > 0:
            Caminho = f"{CaminhoDaPasta}/{formatarNomeDoArquivo(AnoDeApuração, MesApuração, formatarNumeroEmpresa(linha[0].value))}"
            set_icms.set_values(Caminho, linha[1].value, [int(MesApuração), int(AnoDeApuração)])
            
if SelecionarFunção == "GerarProsft":
    GerarArquivoProsoft()
elif SelecionarFunção == "Importar":
    ImportarArquivoSedif()
elif SelecionarFunção == "AdicionarST":
    fix_txt.fix_file(CaminhoDaPasta)
    AdicionarST()
else:
    print("Não foi possível realizar a operação")

